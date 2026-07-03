from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def extract_text(
    file_path: str,
    max_sheets: int = 3,
    max_rows: int = 80,
    max_columns: int = 12,
) -> str:
    path = Path(str(file_path or '').strip()).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f'Excel 文件不存在：{path}')

    sheet_limit = max(1, min(int(max_sheets or 3), 20))
    row_limit = max(1, min(int(max_rows or 80), 500))
    column_limit = max(1, min(int(max_columns or 12), 100))

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sections: list[str] = []
        for worksheet in workbook.worksheets[:sheet_limit]:
            rendered_rows: list[str] = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row or row_limit, row_limit),
                min_col=1,
                max_col=min(worksheet.max_column or column_limit, column_limit),
                values_only=True,
            ):
                values = [str(cell).strip() if cell is not None else '' for cell in row]
                if not any(values):
                    continue
                rendered_rows.append(' | '.join(values))
                if len(rendered_rows) >= row_limit:
                    break
            if rendered_rows:
                sections.append(f'工作表：{worksheet.title}\n' + '\n'.join(rendered_rows))
        return '\n\n'.join(sections).strip()
    finally:
        workbook.close()
